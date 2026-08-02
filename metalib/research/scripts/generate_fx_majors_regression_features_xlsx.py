"""
Features dataframe for FX Majors: HTF (4h) close, CFTC COT net-noncommercial
positioning, and 1-week-forward log return, per pair -- for exploring what
might improve on the Regression Gate. Note: the Regression Gate itself is a
rule-based OLS-slope/R^2 filter on price level, it doesn't have a fitted
forward-return target today -- the fwd_ret_1w_* columns here represent what
a *predictive* regression would be fit against, for future modeling work.

COT is weekly and forward-filled onto the 4h grid with a release-lag shift
(COT data as-of Tuesday is released the following Friday) so the alignment
stays causal -- no COT value is visible before it was actually public.

Usage:
    PYTHONPATH=. "<adonys python>" metalib/research/scripts/generate_fx_majors_regression_features_xlsx.py
"""
import os
import warnings

warnings.filterwarnings("ignore")

from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from metalib.metafvg_backtest import (
    HTF_RESAMPLE_OPTIONS,
    LTF_TIMEFRAME_OPTIONS,
    connect_mt5,
    fetch_ltf_htf_candles,
)

REPORTS_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "reports")
OUT_XLSX = os.path.join(REPORTS_DIR, "fx_majors_regression_features.xlsx")

FX_MAJORS = ["AUDUSD", "EURUSD", "GBPUSD", "NZDUSD", "USDCAD", "USDCHF", "USDJPY"]

# CFTC legacy_fut market-name prefixes per currency, and whether the pair's
# own direction (long pair = pair currency strengthens vs USD) matches the
# futures contract's own direction (long future = that currency strengthens
# vs USD). USDCAD/USDCHF/USDJPY are quoted USD-base, so "long CAD/CHF/JPY
# future" is actually *bearish* the pair -- sign flipped so every cot_<PAIR>
# column consistently means "net speculative position in the direction of
# the quoted pair strengthening."
COT_MARKET_PREFIXES = {
    "EURUSD": (["EURO FX"], 1),
    "GBPUSD": (["BRITISH POUND", "BRITISH POUND STERLING"], 1),  # CFTC renamed "STERLING"->"" in 2022
    "AUDUSD": (["AUSTRALIAN DOLLAR"], 1),
    "NZDUSD": (["NEW ZEALAND DOLLAR", "NZ DOLLAR"], 1),  # CFTC renamed the contract mid-history
    "USDCAD": (["CANADIAN DOLLAR"], -1),
    "USDCHF": (["SWISS FRANC"], -1),
    "USDJPY": (["JAPANESE YEN"], -1),
}

BACKTEST_END = datetime.utcnow()
BACKTEST_START = BACKTEST_END - pd.Timedelta(days=365 * 5)
FORWARD_HORIZON = pd.Timedelta(days=7)
COT_RELEASE_LAG_DAYS = 3


def _normalize_col_name(value: str) -> str:
    """Same normalization the existing pull_cot_euro_positioning_features
    (metalib/mtofx_daily_data.py) relies on: raw CFTC columns look like
    'Noncommercial Positions-Long (All)' -- lowercase, non-alnum -> '_',
    collapse repeats."""
    cleaned = "".join(ch.lower() if ch.isalnum() else "_" for ch in str(value).strip())
    while "__" in cleaned:
        cleaned = cleaned.replace("__", "_")
    return cleaned.strip("_")


def pull_cot_net_positioning(
    market_prefixes: list[str],
    sign: int,
    start: pd.Timestamp,
    end: pd.Timestamp,
    report_type: str = "legacy_fut",
    release_lag_days: int = COT_RELEASE_LAG_DAYS,
) -> pd.Series:
    """
    Weekly net noncommercial (speculator) positioning = long - short, for
    whichever CFTC market name(s) start with one of market_prefixes (prefix
    match, not substring -- substring would false-positive-match cross-rate
    contracts like "EURO FX/BRITISH POUND XRATE"). Index shifted forward by
    release_lag_days so it's causal (COT as-of Tuesday, released the
    following Friday).
    """
    import cot_reports as cot

    years = range(start.year, end.year + 1)
    yearly = []
    for year in years:
        try:
            yearly.append(cot.cot_year(year=year, cot_report_type=report_type))
        except Exception:
            continue
    if not yearly:
        raise RuntimeError(f"No COT data fetched for years {list(years)}")

    raw = pd.concat(yearly, ignore_index=True)
    raw.columns = [_normalize_col_name(c) for c in raw.columns]

    def first_col(candidates):
        for c in candidates:
            if c in raw.columns:
                return c
        raise KeyError(f"None of {candidates} found in columns {list(raw.columns)[:15]}...")

    market_col = first_col(["market_and_exchange_names"])
    date_col = first_col(["as_of_date_in_form_yyyy_mm_dd", "report_date_as_yyyy_mm_dd"])
    long_col = first_col(["noncommercial_positions_long_all"])
    short_col = first_col(["noncommercial_positions_short_all"])

    market = raw[market_col].astype(str).str.upper()
    mask = pd.Series(False, index=raw.index)
    for prefix in market_prefixes:
        # Require " -" right after the prefix (the real contract's name is
        # "<CURRENCY> - CHICAGO MERCANTILE EXCHANGE"), not just a bare
        # prefix -- "EURO FX" alone also prefix-matches the unrelated
        # cross-rate contract "EURO FX/BRITISH POUND XRATE - ...", which has
        # a completely different position scale and silently corrupts the
        # series if it slips through (confirmed empirically: one bogus
        # 119,476 spike between two ~-2000-4000 EURUSD weeks).
        mask |= market.str.startswith(prefix.upper() + " -")
    sub = raw[mask].copy()
    if sub.empty:
        raise RuntimeError(f"No COT rows matched prefixes {market_prefixes}")

    sub["date"] = pd.to_datetime(sub[date_col], errors="coerce").dt.normalize()
    long_v = pd.to_numeric(sub[long_col].astype(str).str.replace(",", "", regex=False), errors="coerce")
    short_v = pd.to_numeric(sub[short_col].astype(str).str.replace(",", "", regex=False), errors="coerce")
    sub["net"] = sign * (long_v - short_v)
    sub = sub.dropna(subset=["date", "net"]).sort_values("date")

    weekly = sub.groupby("date")["net"].last()
    weekly.index = weekly.index + pd.Timedelta(days=release_lag_days)
    weekly = weekly.sort_index()
    return weekly.loc[(weekly.index >= start) & (weekly.index <= end + pd.Timedelta(days=release_lag_days + 14))]


def forward_log_return(close: pd.Series, horizon: pd.Timedelta) -> pd.Series:
    """Calendar-time forward return (not a fixed bar-count shift) -- robust to
    weekend gaps in the HTF grid, so "1 week ahead" means 1 real week, not
    "N bars ahead" where N silently drifts depending on how many weekend
    gaps happened to fall in that window."""
    idx = close.index
    target_times = idx + horizon
    pos = idx.searchsorted(target_times, side="left")
    fwd = np.full(len(close), np.nan)
    valid = pos < len(idx)
    fwd[valid] = close.values[pos[valid]]
    fwd_close = pd.Series(fwd, index=idx)
    with np.errstate(divide="ignore", invalid="ignore"):
        return np.log(fwd_close) - np.log(close)


def build_features() -> pd.DataFrame:
    connect_mt5()

    closes: dict[str, pd.Series] = {}
    fwd_rets: dict[str, pd.Series] = {}
    for symbol in FX_MAJORS:
        print(f"fetching {symbol}...", flush=True)
        _, htf_df = fetch_ltf_htf_candles(
            symbol, BACKTEST_START, BACKTEST_END,
            ltf_timeframe=LTF_TIMEFRAME_OPTIONS["M15"], htf_resample_rule=HTF_RESAMPLE_OPTIONS["4h"],
        )
        close = htf_df["close"].dropna()
        closes[symbol] = close
        fwd_rets[symbol] = forward_log_return(close, FORWARD_HORIZON)

    master_index = sorted(set().union(*[c.index for c in closes.values()]))
    master_index = pd.DatetimeIndex(master_index)

    cots: dict[str, pd.Series] = {}
    for symbol in FX_MAJORS:
        prefixes, sign = COT_MARKET_PREFIXES[symbol]
        print(f"fetching COT for {symbol} ({prefixes}, sign={sign})...", flush=True)
        weekly = pull_cot_net_positioning(prefixes, sign, BACKTEST_START, BACKTEST_END)
        cots[symbol] = weekly

    df = pd.DataFrame(index=master_index)
    for symbol in FX_MAJORS:
        df[f"close_{symbol}"] = closes[symbol].reindex(master_index, method="ffill")
    for symbol in FX_MAJORS:
        weekly = cots[symbol].reindex(cots[symbol].index.union(master_index)).ffill().reindex(master_index)
        df[f"cot_{symbol}"] = weekly
    for symbol in FX_MAJORS:
        df[f"fwd_ret_1w_{symbol}"] = fwd_rets[symbol].reindex(master_index, method="ffill")

    df.index.name = "time"
    return df


def write_xlsx(df: pd.DataFrame, out_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "features"

    headers = ["time"] + list(df.columns)
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4A9EFF")
        cell.alignment = Alignment(horizontal="center")

    close_cols = [c for c in df.columns if c.startswith("close_")]
    cot_cols = [c for c in df.columns if c.startswith("cot_")]
    fwd_cols = [c for c in df.columns if c.startswith("fwd_ret_")]

    for row_i, (ts, row) in enumerate(df.iterrows(), start=2):
        values = [ts.to_pydatetime()] + [row[c] for c in df.columns]
        ws.append(values)
        ws.cell(row=row_i, column=1).number_format = "yyyy-mm-dd hh:mm"
        for c, col_name in enumerate(df.columns, start=2):
            cell = ws.cell(row=row_i, column=c)
            if col_name in close_cols:
                cell.number_format = "0.00000"
            elif col_name in cot_cols:
                cell.number_format = "#,##0"
            elif col_name in fwd_cols:
                cell.number_format = "0.0000%"

    for c, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(c)].width = max(11, len(h) + 2)
    ws.freeze_panes = "B2"

    wb.save(out_path)


def main():
    df = build_features()
    print(f"assembled dataframe: {df.shape[0]} rows x {df.shape[1]} cols, "
          f"{df.index.min()} -> {df.index.max()}", flush=True)
    write_xlsx(df, OUT_XLSX)
    print(f"written to: {OUT_XLSX}", flush=True)


if __name__ == "__main__":
    main()
