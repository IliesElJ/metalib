"""
One-off: per-bucket performance table for Baseline (unmodified MetaFVG) vs.
Regression Gate vs. Spearman Gate at its three tested thresholds (0.5 loose,
0.82, 0.85 tight), pulled from the M15/4h A/B sweep cache. Requested as an
Excel table rather than the usual PDF report -- ad hoc, not wired into the
main report pipeline.

Usage:
    PYTHONPATH=. "<adonys python>" metalib/research/scripts/generate_bucket_spearman_xlsx.py --bucket "FX Majors"
    PYTHONPATH=. "<adonys python>" metalib/research/scripts/generate_bucket_spearman_xlsx.py --bucket Indices
"""
import argparse
import os
import pickle
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from metafvg_ab_universe import UNIVERSE  # noqa: E402

DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data")
REPORTS_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "reports")
CACHE_PATH = os.path.join(DATA_DIR, "metafvg_ab_sweep_data_m15_4h.pkl")

CONFIGS = ["Baseline", "Regression Gate", "Spearman Gate", "Spearman Gate (0.82)", "Spearman Gate (tight)"]
CONFIG_LABELS = {
    "Baseline": "Baseline (no gate)",
    "Regression Gate": "Regression Gate (OLS R2>=0.5)",
    "Spearman Gate": "Spearman (rho>=0.5)",
    "Spearman Gate (0.82)": "Spearman (rho>=0.82)",
    "Spearman Gate (tight)": "Spearman (rho>=0.85)",
}

COLUMNS = [
    ("symbol", "Symbol", "text"),
    ("config", "Config", "text"),
    ("n_closed", "Closed Trades", "int"),
    ("win_rate_pct", "Win Rate %", "pct1"),
    ("avg_winner_r", "Avg Winner (R)", "num2"),
    ("profit_factor", "Profit Factor", "num2"),
    ("breakeven_wr_pct", "Breakeven WR %", "pct1"),
    ("edge_pp", "Edge (pp)", "num2"),
    ("sharpe", "Sharpe (vbt)", "num3"),
    ("qs_sortino", "Sortino", "num3"),
    ("qs_calmar", "Calmar", "num3"),
    ("total_return_pct", "Total Return %", "pct1_raw"),
    ("max_dd_pct", "Max DD %", "pct1_raw"),
    ("qs_kelly", "Kelly", "num3"),
]

NUM_FORMATS = {
    "int": "0",
    "pct1": "0.0",
    "pct1_raw": "0.0%",
    "num2": "0.00",
    "num3": "0.000",
}

BG = "0F1117"
PANEL = "1A1D27"
HEADER_BLUE = "4A9EFF"
TEXT = "E0E0E0"
GREEN = "00C853"
RED = "FF3D3D"


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER_BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def write_sheet_by_symbol(wb, data, symbols):
    ws = wb.active
    ws.title = "By Symbol"
    headers = [label for _, label, _ in COLUMNS]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    row_i = 2
    for symbol in symbols:
        for cfg in CONFIGS:
            rec = data.get((symbol, cfg))
            if rec is None:
                continue
            row = []
            for key, _, kind in COLUMNS:
                if key == "config":
                    row.append(CONFIG_LABELS[cfg])
                elif key in ("total_return_pct", "max_dd_pct"):
                    row.append(rec.get(key, float("nan")) / 100.0)
                else:
                    row.append(rec.get(key, float("nan")))
            ws.append(row)
            for c, (key, _, kind) in enumerate(COLUMNS, start=1):
                cell = ws.cell(row=row_i, column=c)
                if kind in NUM_FORMATS:
                    cell.number_format = NUM_FORMATS[kind]
                cell.font = Font(color=TEXT)
                cell.fill = PatternFill("solid", fgColor=PANEL if row_i % 2 == 0 else BG)
                if key == "sharpe":
                    val = rec.get("sharpe")
                    if val is not None and val == val:
                        cell.font = Font(color=GREEN if val > 0 else RED)
            row_i += 1

    for c, (_, label, _) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(c)].width = max(12, len(label) + 2)
    ws.freeze_panes = "A2"


def write_sheet_summary(wb, data, symbols):
    ws = wb.create_sheet(f"Summary (avg across {len(symbols)} symbols)")
    headers = ["Config", "Avg Closed Trades", "Avg Win Rate %", "Avg Sharpe (vbt)", "Median Sharpe (vbt)",
               "Avg Sortino", "Avg Total Return %", "Avg Max DD %"]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    import statistics as stats

    row_i = 2
    for cfg in CONFIGS:
        recs = [data[(s, cfg)] for s in symbols if (s, cfg) in data]
        sharpes = [r["sharpe"] for r in recs if r.get("sharpe") == r.get("sharpe")]
        row = [
            CONFIG_LABELS[cfg],
            sum(r["n_closed"] for r in recs) / len(recs) if recs else float("nan"),
            sum(r["win_rate_pct"] for r in recs) / len(recs) if recs else float("nan"),
            sum(sharpes) / len(sharpes) if sharpes else float("nan"),
            stats.median(sharpes) if sharpes else float("nan"),
            sum(r["qs_sortino"] for r in recs) / len(recs) if recs else float("nan"),
            sum(r["total_return_pct"] for r in recs) / len(recs) / 100.0 if recs else float("nan"),
            sum(r["max_dd_pct"] for r in recs) / len(recs) / 100.0 if recs else float("nan"),
        ]
        ws.append(row)
        fmts = ["", "0.0", "0.0", "0.000", "0.000", "0.000", "0.0%", "0.0%"]
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row_i, column=c)
            if fmts[c - 1]:
                cell.number_format = fmts[c - 1]
            cell.font = Font(color=TEXT)
            cell.fill = PatternFill("solid", fgColor=PANEL if row_i % 2 == 0 else BG)
        row_i += 1

    for c, label in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(c)].width = max(14, len(label) + 2)
    ws.freeze_panes = "A2"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--bucket", required=True, help=f"one of {list(UNIVERSE.keys())}")
    args = parser.parse_args()

    if args.bucket not in UNIVERSE:
        raise SystemExit(f"Unknown bucket {args.bucket!r}. Options: {list(UNIVERSE.keys())}")
    symbols = UNIVERSE[args.bucket]

    with open(CACHE_PATH, "rb") as f:
        data = pickle.load(f)

    slug = args.bucket.lower().replace(" ", "_")
    out_path = os.path.join(REPORTS_DIR, f"metafvg_{slug}_spearman_comparison.xlsx")

    wb = Workbook()
    write_sheet_by_symbol(wb, data, symbols)
    write_sheet_summary(wb, data, symbols)
    wb.save(out_path)
    print(f"Written to: {out_path}")


if __name__ == "__main__":
    main()
